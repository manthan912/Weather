from pyowm.owm import OWM
import openpyxl as op

z = op.Workbook()
s = z['Sheet']
s.title = "Weather"
z.create_sheet(title="Cities")
c = z['Cities']
c.merge_cells("A1:B1")
c.cell(row=1,column=1).value = "CITIES"
c.merge_cells("A2:B2")
c.cell(row=2,column=1).value = "Mumbai"
c.merge_cells("A3:B3")
c.cell(row=3,column=1).value = "New York"
c.merge_cells("A4:B4")
c.cell(row=4,column=1).value = "London"
c.merge_cells("A5:B5")
c.cell(row=5,column=1).value = "Paris"
s.merge_cells("A1:B1")
s.cell(row=1,column=1).value = "City"
s.cell(row=1,column=3).value = "temp"
s.cell(row=1,column=4).value = "Temp_max"
s.cell(row=1,column=5).value = "Temp_min"
for i in range(2,6):
    city =c.cell(row=i,column=1).value
    owm = OWM("25ae03a24d9b87c8eec53ff073b973b1")
    mgr = owm.weather_manager()
    observation = mgr.weather_at_place(city)  # the observation object is a box containing a weather object
    weather = observation.weather
    temfar = weather.temperature('celsius')
    s.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
    s.cell(row=i, column=1).value = city
    s.cell(row=i, column=3).value = temfar['temp']
    s.cell(row=i, column=4).value = temfar["temp_max"]
    s.cell(row=i, column=5).value = temfar['temp_min']


z.save('C:\\Users\\MANTHAN\\Desktop\\Weather.xlsx')
